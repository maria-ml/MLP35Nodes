# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 17:35:31 2023

@author: maria
"""
from openpyxl import load_workbook
import numpy as np
import openpyxl

# CLASIFICADOR HEALTHY/NOT HEALTHY PARA FRAMES

# EL FICHERO QUE SE LEE ES EL EXCEL

rutaleer = 'C:/Users/maria/OneDrive - Fundación Universitaria San Pablo CEU/CEU/Cursos/TERCERO/SEGUNDO CUATRIMESTRE/PROYECTOS II/WEKA/Resultados/Excels/Finales/'
nombre = 'total35_data' # !!!!!!!!!!!!!!!!!

archivoleer = rutaleer + nombre 

# LEER FICHEROS
workbook = load_workbook(archivoleer + '.xlsx')

# Seleccionar la hoja de trabajo
sheet = workbook.active

# Obtener los valores de las cinco columnas
frame = [cell.value for cell in sheet['A']]
actual = [cell.value.hour for cell in sheet['B']]
predicted = [cell.value.hour for cell in sheet['C']]
error = [cell.value for cell in sheet['D']]
prediction = [cell.value for cell in sheet['E']]

# para clasificar en healthy no healthy solo tenemos que fijarnos en la columna de predicted
pred_diagnosis = []
for pred in predicted:
    if pred == 1 :
        pred_diagnosis.append('Healthy')
    else:
        pred_diagnosis.append('Not healthy')
        
# print(pred_diagnosis)

true_diagnosis = []
for true in actual:
    if true == 1 :
        true_diagnosis.append('Healthy')
    else:
        true_diagnosis.append('Not healthy')
        

# # Definir los datos
# data = [frame, true_diagnosis, pred_diagnosis]

# # Iterar sobre los datos y escribir cada columna
# for row in data:
#     sheet.append(row)

# # Guardar el archivo
# file = archivoleer + 'H_NH' + '.xlsx'
# workbook.save(file)

workbook = openpyxl.Workbook()

# Seleccionar la hoja activa
worksheet = workbook.active

for i in range(1, len(frame)+1):
    worksheet.cell(row=i, column=1, value=frame[i-1])
    worksheet.cell(row=i, column=2, value=true_diagnosis[i-1])
    worksheet.cell(row=i, column=3, value=pred_diagnosis[i-1])

# Guardar el archivo Excel
workbook.save(archivoleer + 'H_NH' + '.xlsx')






