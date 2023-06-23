# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 09:18:10 2023

@author: maria
"""


from sklearn.metrics import confusion_matrix
from openpyxl import load_workbook
import numpy as np

rutaleer = 'C:/Users/maria/OneDrive - Fundación Universitaria San Pablo CEU/CEU/Cursos/TERCERO/SEGUNDO CUATRIMESTRE/PROYECTOS II/WEKA/Resultados/Excels/Finales/'
# ruta = 'C:/Users/maria/Documents/CEU/weka/Resultados de verdad/'
ruta = 'C:/Users/maria/Documents/CEU/weka/'
rutaguardar = 'C:/Users/maria/OneDrive - Fundación Universitaria San Pablo CEU/CEU/Cursos/TERCERO/SEGUNDO CUATRIMESTRE/PROYECTOS II/WEKA/Resultados/Matrices/'


# EL FICHERO QUE SE LEE ES EL EXCEL
# ESTO ES PARA LA CLASIFICACIÓN POR FRAMES

nombre = 'total35_Files_H_NH_' #!!!!!!!!!!!

archivoleer = rutaleer + nombre + 'data'


# def primer_num(cadena):
#     primer_numero = ''
#     for caracter in cadena:
#         if caracter.isdigit():
#             primer_numero = caracter
#             break
#     return primer_numero


# LEER FICHEROS
workbook = load_workbook(archivoleer + '.xlsx')

# Seleccionar la hoja de trabajo
sheet = workbook.active

# Obtener los valores de las cinco columnas
frame = [cell.value for cell in sheet['A']]
# actual = [cell.value.hour for cell in sheet['B']]
actual = [cell.value for cell in sheet['B']] #!!!!!!!!!!!!!!!!!!!
# predicted = [cell.value.hour for cell in sheet['C']]
predicted = [cell.value for cell in sheet['C']] #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
error = [cell.value for cell in sheet['D']]
prediction = [cell.value for cell in sheet['E']]

# Imprimir los valores de las columnas
# print(columna1)
# print(columna2)
# print(columna3)
# print(columna4)
# print(columna5)

confusionMatrix = confusion_matrix(actual, predicted)
print(confusionMatrix)

# GUARDAR FICHEROS
fichero1 = rutaguardar + 'm' + nombre + '.txt'
np.savetxt(fichero1, confusionMatrix, fmt='%d') 










# for col in dataframe1.iter_cols(1, dataframe1.max_column):
#     for row in range(0, dataframe1.max_row):
        
#         valor = col[row].value
#         valor = str(valor)
#         if valor.isalpha():
#             addnumber(diagnosis, valor)
            
#         else:
#             nombre.append(valor)


#-------------------------------------------------------
# with open(archivo, "r") as archivo:
#     lineas = archivo.readlines()

# c1 = []
# actual = []
# predicted = []
# c4 = []
# c5 = []


# print(len(lineas))
# for linea in lineas:
#     # print(linea)
#     columnas = re.findall(r"\S+", linea)
#     # print(len(columnas))
#     for i in range (0, len(columnas)):
#         columna = columnas[i]
#     # for columna in columnas:
#         if not any(caracter.isdigit() or caracter == '+' for caracter in columna):
#             continue
#         print (i)
#         print(columna)




    # print('siguiente linea')
# print(len(lineas))

# for linea in lineas:
#     # print(linea)
#     columnas = re.findall(r"\S+", linea)
#     if not any(caracter.isdigit() or caracter == '+' for caracter in linea):
#         continue
#     if any(caracter == '+' for caracter in linea):
#         v1 = columnas[5]
#         print(v1)
#         c1.append(v1)
#         actual.append(columnas[13])
#         predicted.append(columnas[21])
#         c4.append(columnas[24])
#         c5.append(columnas[27])
#     else:
#         v1 = columnas[4-5]
#         # print(v1)
#         c1.append(v1)
#         actual.append(columnas[12-13])
#         predicted.append(columnas[20-21])
#         c4.append(' ')
#         c5.append(columnas[27-28])


# print(c1)
# print(actual)
# print(predicted)
# print(c4)
# print(c5)


#-------------------------------------------------------




# with open(archivo, 'r') as archivo:
#     lineas = archivo.readlines()

# print(len(lineas))

# for linea in lineas:
#     # Dividir la línea en columnas utilizando varios espacios como separadores
#     columnas = linea.split(' ')

#     # Acceder a los valores de cada columna
#     columna1 = columnas[0]
#     columna2 = columnas[1]
#     columna3 = columnas[2]
#     print(columna1)




#-------------------------------------------------------


    # # columna1 = linea[0:9]  # Rebanado para extraer la columna 1 (ancho: 10)
    # columna1 = linea[0:18]
    # # print(columna1)
    # # v = columna1.strip()
    # # print(v)
    # # c1.append(v)
    # columna2 = linea[19:40]  # Rebanado para extraer la columna 2 (ancho: 10)
    # # print(columna2)
    # if columna2.isspace():
    #     print('hola')
    # v = primer_num(columna2)
    # # print(v)
    # # print('hola')
    # actual.append(v)
    # # c2.append(v)
    # # print(c2)
    # columna3 = linea[41:62]  # Rebanado para extraer la columna 3 (ancho: 10)
    # # print(columna3)
    # # v = columna3.strip()
    # # c3.append(v)
    # v = primer_num(columna3)
    # # print(v)
    # predicted.append(v)
    # columna4 = linea[63:76]  # Rebanado para extraer la columna 4 (ancho: 10)
    # # print(columna4)
    # # v = columna4.strip()
    # # c4.append(v)
    # columna5 = linea[77:86]  # Rebanado para extraer la columna 5 (ancho: 10)
    # # print(columna5)
    # # v = columna5.strip()
    # # c5.append(v)

# print('Actual values:')
# print(actual)
# print('Predicted values')
# print(predicted)
# matrix = [actual, predicted]
# print(matrix)

# confusionMatrix = confusion_matrix(actual, predicted)
# print(confusionMatrix)

